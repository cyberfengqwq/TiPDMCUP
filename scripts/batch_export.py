#!/usr/bin/env python3
"""
批量导出赛题答案到 result_2.xlsx / result_3.xlsx

用法:
    python scripts/batch_export.py 2          # 只导出任务二
    python scripts/batch_export.py 3          # 只导出任务三
    python scripts/batch_export.py all        # 全部导出
    python scripts/batch_export.py 2 --data-dir /path/to/附件 --output-dir ./output
"""

import argparse
import json
import sys
import time
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from core.agent.pipeline import Agent
from core.services.vllm_service import TransformersLLM

ANALYSIS_MODEL = "/home/qwq/TiPDMCUP/models/Qwen2.5-7B-Coder-Instruct"
SQL_MODEL = "/home/qwq/TiPDMCUP/models/Qwen2.5-7B-N2SQL"

BATCH_USER_ID = "batch_export"
BATCH_COMPANY_ID = "batch"
# 每次运行用唯一时间戳，避免跨次运行的 chat 历史污染
_RUN_TS = int(time.time())

CHART_TYPE_ZH = {
    "line": "折线图",
    "bar": "柱状图",
    "pie": "饼图",
    "table": "表格",
    "none": "无",
}


def load_llms() -> tuple[TransformersLLM, TransformersLLM]:
    # 只创建对象，不预加载——pipeline 每次调用时按需加载/卸载，避免两模型同时占 GPU
    sql_llm = TransformersLLM(
        _modelpath=SQL_MODEL,
        _temperature=0.2,
        _top_p=0.9,
        _max_tokens=512,
    )
    analysis_llm = TransformersLLM(
        _modelpath=ANALYSIS_MODEL,
        _temperature=0.7,
        _top_p=0.8,
        _max_tokens=1024,
    )
    print("LLM 对象已创建（模型将在首次调用时按需加载）\n")
    return sql_llm, analysis_llm


def read_questions(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if row[0] is None:
            continue
        rows.append(
            {
                "problem_id": str(row[0]),
                "question_type": str(row[1] or ""),
                "questions_raw": str(row[2] or ""),
            }
        )
    return rows


def run_problem(
    sql_llm: TransformersLLM,
    analysis_llm: TransformersLLM,
    problem_id: str,
    questions: list[str],
    task: int,
    question_type: str,
    questions_raw: str,
    detailed_results: list[dict] | None,
    output_json_path: Path | None,
) -> dict:
    """为一道题新建 Agent，顺序跑完所有对话轮次，返回汇总结果。"""
    # task=2：用轻量规则门卫（无需 ML 模型），保留多轮追问过程
    # task=3：模糊/多意图查询直接进分解，跳过门卫
    agent = Agent(
        user_id=BATCH_USER_ID,
        company_id=BATCH_COMPANY_ID,
        chat_id=f"batch_{problem_id}_{_RUN_TS}",
        sql_llm=sql_llm,
        analysis_llm=analysis_llm,
        skip_gatekeeper=(task == 3),
        use_lite_gatekeeper=(task == 2),
    )

    qa_pairs: list[dict] = []
    sqls: list[str] = []
    chart_types: list[str] = []
    rounds: list[dict] = []

    for idx, q in enumerate(questions, start=1):
        result = agent.run(q, problem_id=problem_id, task=task)
        a = result.get("A", {})

        content = a.get("content", "")
        image = a.get("image", [])
        sql = a.get("sql", "")
        ct = a.get("chart_type", "none")
        references = a.get("references", [])

        if sql and not sql.startswith("--"):
            sqls.append(sql)
        if ct != "none":
            chart_types.append(ct)

        entry: dict = {"Q": q, "A": {"content": content, "image": image}}
        if references:
            entry["A"]["references"] = references
        qa_pairs.append(entry)

        rounds.append({
            "round": idx,
            "question": q,
            "intent_info": a.get("intent_info", {}),
            "rag_context": a.get("rag_context", {}),
            "sql": sql,
            "sql_corrected": a.get("sql_corrected", False),
            "sql_result": a.get("sql_result", []),
            "analysis": content,
            "image": image,
        })

    # 每道题一条记录，所有轮次嵌套在 rounds 里
    problem_record = {
        "problem_id": problem_id,
        "task": task,
        "question_type": question_type,
        "questions_raw": questions_raw,
        "qa_pairs": qa_pairs,
        "sqls": sqls,
        "rounds": rounds,
    }
    if detailed_results is not None:
        detailed_results.append(problem_record)
        if output_json_path is not None:
            write_batch_json(detailed_results, output_json_path)

    return {
        "qa_pairs": qa_pairs,
        "sqls": sqls,
        "chart_types": chart_types,
    }


def _write_header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)


def _set_col_widths(ws, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")


def write_batch_json(records: list[dict], output_path: Path) -> None:
    def _default(o):
        if o.__class__.__name__ == "Decimal":
            return float(o)
        raise TypeError(
            f"Object of type {o.__class__.__name__} is not JSON serializable"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2, default=_default)


def export_task2(
    questions_xlsx: Path,
    output_path: Path,
    sql_llm: TransformersLLM,
    analysis_llm: TransformersLLM,
    detailed_results: list[dict],
    output_json_path: Path,
) -> None:
    rows = read_questions(questions_xlsx)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_header_row(ws, ["编号", "问题", "SQL查询语句", "图形格式", "回答"])

    for row in rows:
        problem_id = row["problem_id"]
        questions_raw = row["questions_raw"]

        try:
            q_list = json.loads(questions_raw)
            if not isinstance(q_list, list):
                raise ValueError("questions_raw is not a list")
            questions = [
                item.get("Q") if isinstance(item, dict) else str(item)
                for item in q_list
            ]
        except Exception as e:
            print(f"[{problem_id}] 解析问题失败: {e}")
            detailed_results.append({
                "problem_id": problem_id,
                "task": 2,
                "question_type": row["question_type"],
                "questions_raw": questions_raw,
                "qa_pairs": [],
                "sqls": [],
                "rounds": [{"round": 0, "question": "", "error": str(e)}],
            })
            write_batch_json(detailed_results, output_json_path)
            continue

        print(f"[{problem_id}] 开始处理 ({len(questions)} 轮)...")
        result = run_problem(
            sql_llm,
            analysis_llm,
            problem_id,
            questions,
            task=2,
            question_type=row["question_type"],
            questions_raw=questions_raw,
            detailed_results=detailed_results,
            output_json_path=output_json_path,
        )

        sql_str = "\n---\n".join(result["sqls"]) if result["sqls"] else "无"
        chart_str = (
            "；".join(CHART_TYPE_ZH.get(ct, ct) for ct in result["chart_types"])
            if result["chart_types"]
            else "无"
        )
        answer_json = json.dumps(result["qa_pairs"], ensure_ascii=False, indent=None)

        ws.append([problem_id, questions_raw, sql_str, chart_str, answer_json])
        print(f"[{problem_id}] 完成")

    _set_col_widths(ws, {"A": 10, "B": 30, "C": 40, "D": 12, "E": 80})
    wb.save(output_path)
    print(f"\n✓ 已保存: {output_path}")


def export_task3(
    questions_xlsx: Path,
    output_path: Path,
    sql_llm: TransformersLLM,
    analysis_llm: TransformersLLM,
    detailed_results: list[dict],
    output_json_path: Path,
) -> None:
    rows = read_questions(questions_xlsx)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _write_header_row(ws, ["编号", "问题", "SQL查询语句", "回答"])

    for row in rows:
        problem_id = row["problem_id"]
        questions_raw = row["questions_raw"]

        try:
            q_list = json.loads(questions_raw)
            if not isinstance(q_list, list):
                raise ValueError("questions_raw is not a list")
            questions = [
                item.get("Q") if isinstance(item, dict) else str(item)
                for item in q_list
            ]
        except Exception as e:
            print(f"[{problem_id}] 解析问题失败: {e}")
            detailed_results.append({
                "problem_id": problem_id,
                "task": 3,
                "question_type": row["question_type"],
                "questions_raw": questions_raw,
                "qa_pairs": [],
                "sqls": [],
                "rounds": [{"round": 0, "question": "", "error": str(e)}],
            })
            write_batch_json(detailed_results, output_json_path)
            continue

        print(f"[{problem_id}] 开始处理 ({len(questions)} 轮)...")
        result = run_problem(
            sql_llm,
            analysis_llm,
            problem_id,
            questions,
            task=3,
            question_type=row["question_type"],
            questions_raw=questions_raw,
            detailed_results=detailed_results,
            output_json_path=output_json_path,
        )

        sql_str = "\n---\n".join(result["sqls"]) if result["sqls"] else "无"
        answer_json = json.dumps(result["qa_pairs"], ensure_ascii=False, indent=None)

        ws.append([problem_id, questions_raw, sql_str, answer_json])
        print(f"[{problem_id}] 完成")

    _set_col_widths(ws, {"A": 10, "B": 30, "C": 40, "D": 80})
    wb.save(output_path)
    print(f"\n✓ 已保存: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="批量导出赛题答案")
    parser.add_argument("task", choices=["2", "3", "all"], help="导出任务二、三或全部")
    parser.add_argument(
        "--data-dir",
        default="示例数据",
        help="附件所在目录（默认: 示例数据）",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="输出目录（默认: 当前目录）",
    )
    args = parser.parse_args()

    data_dir = Path(args.data_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    sql_llm, analysis_llm = load_llms()
    detailed_results: list[dict] = []
    output_json_path = output_dir / "batch_results.json"
    write_batch_json(detailed_results, output_json_path)

    if args.task in ("2", "all"):
        export_task2(
            questions_xlsx=data_dir / "附件4：问题汇总.xlsx",
            output_path=output_dir / "result_2.xlsx",
            sql_llm=sql_llm,
            analysis_llm=analysis_llm,
            detailed_results=detailed_results,
            output_json_path=output_json_path,
        )

    if args.task in ("3", "all"):
        export_task3(
            questions_xlsx=data_dir / "附件6：问题汇总.xlsx",
            output_path=output_dir / "result_3.xlsx",
            sql_llm=sql_llm,
            analysis_llm=analysis_llm,
            detailed_results=detailed_results,
            output_json_path=output_json_path,
        )

    if detailed_results:
        write_batch_json(detailed_results, output_dir / "batch_results.json")


if __name__ == "__main__":
    main()
